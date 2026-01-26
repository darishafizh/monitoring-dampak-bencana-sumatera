
import openpyxl
import json
import re

def clean_value(val):
    if val is None:
        return 0
    if isinstance(val, str):
        val = val.strip().replace(',', '').replace('.', '')
        if val == '-' or val == '':
            return 0
        try:
            return float(val)
        except:
            return 0 # If formula or text, ignore for now (or handle formulas if possible)
    return val

def clean_text(val):
    if val is None:
        return ""
    return str(val).strip()

def run():
    print("Loading workbook...")
    wb = openpyxl.load_workbook('2026.01.23_DataRekap_BantuanPascaBencana_(230126).xlsx', data_only=True)
    ws = wb['Sheet1']
    
    data_map = {} # Key: "Provinsi-Kabupaten", Value: { id, provinsi, kabupaten, items: [], totalBiaya }
    
    current_unit = "Unknown"
    
    # Define Column Blocks mapping
    # (Label, KabCol, VolCol, BiayaCol, IsOnTop)
    # Based on Row 5/6 headers
    blocks = [
        # APBN Reguler
        ("Reguler - Aceh", 9, 10, 11, False),
        ("Reguler - Sumut", 12, 13, 14, False),
        ("Reguler - Sumbar", 15, 16, 17, False),
        # APBN Tambahan (On Top)
        ("Tambahan - Aceh", 21, 22, 23, True),
        ("Tambahan - Sumut", 24, 25, 26, True),
        ("Tambahan - Sumbar", 27, 28, 29, True)
    ]
    
    # Start iterating from data rows
    # Row 7 is where data starts (based on 'BPPSDM' in A7)
    start_row = 7
    end_row = ws.max_row
    
    print(f"Processing rows {start_row} to {end_row}...")
    
    for row_idx in range(start_row, end_row + 1):
        # Check for Unit in Col A (1)
        col1_val = ws.cell(row=row_idx, column=1).value
        if col1_val:
            current_unit = str(col1_val).strip()
            # If current_unit is like "Version:...", skip
            if "Version" in current_unit:
                continue
                
        # Get Activity Name from Col G (7)
        activity_name = ws.cell(row=row_idx, column=7).value
        satuan = ws.cell(row=row_idx, column=8).value
        
        if not activity_name:
            continue
            
        activity_name = clean_text(activity_name)
        satuan = clean_text(satuan)
        
        # Iterate through the 6 blocks
        for block_name, kab_col, vol_col, biaya_col, is_on_top in blocks:
            kab_val = ws.cell(row=row_idx, column=kab_col).value
            
            if kab_val and isinstance(kab_val, str) and len(kab_val) > 2 and "=" not in kab_val:
                # We have a location!
                kabupaten = clean_text(kab_val)
                
                # Determine Province based on block name
                if "Aceh" in block_name:
                    provinsi = "Aceh"
                elif "Sumut" in block_name:
                    provinsi = "Sumatera Utara"
                elif "Sumbar" in block_name:
                    provinsi = "Sumatera Barat"
                else:
                    provinsi = "Unknown"
                
                volume = clean_value(ws.cell(row=row_idx, column=vol_col).value)
                biaya = clean_value(ws.cell(row=row_idx, column=biaya_col).value)
                
                key = f"{provinsi}-{kabupaten}"
                
                if key not in data_map:
                    data_map[key] = {
                        "id": key.lower().replace(" ", "_").replace(".", ""),
                        "provinsi": provinsi,
                        "kabupaten": kabupaten,
                        "items": [],
                        "totalBiaya": 0
                    }
                
                # Add item
                if biaya > 0 or volume > 0:
                    item_entry = {
                        "unit": current_unit,
                        "kegiatan": activity_name,
                        "kategori": "Reguler" if not is_on_top else "Tambahan (On Top)",
                        "satuan": satuan,
                        "volume": volume,
                        "biaya": biaya
                    }
                    data_map[key]["items"].append(item_entry)
                    data_map[key]["totalBiaya"] += biaya

    # Convert map to list and sort
    final_data = list(data_map.values())
    final_data.sort(key=lambda x: (x["provinsi"], x["kabupaten"]))
    
    # Filter out entries with 0 total cost and 0 items (unless needed)
    final_data = [d for d in final_data if d["totalBiaya"] > 0 or len(d["items"]) > 0]

    # Write JS file
    js_content = f"""// Seed data untuk Dashboard Monitoring Dampak Bencana Sumatera
// Generated from Excel: 2026.01.23_DataRekap_BantuanPascaBencana_(230126).xlsx
// Date: {re.sub(r'[^0-9]', '', str(openpyxl.__version__))}

const seedData = {json.dumps(final_data, indent=4)};

// Force load new data on initialization
if (typeof window !== 'undefined') {{
    document.addEventListener('DOMContentLoaded', function() {{
        localStorage.setItem('budidayaData', JSON.stringify(seedData));
        console.log('✅ Data migrated successfully! Total: ' + seedData.length + ' kabupaten entries.');
        
        if (typeof refreshAll === 'function') {{
            refreshAll();
        }}
    }});
}}

// Export for manual use
if (typeof module !== 'undefined' && module.exports) {{
    module.exports = {{ seedData }};
}}
"""
    
    with open('seed-data.js', 'w', encoding='utf-8') as f:
        f.write(js_content)
        
    print(f"Successfully generated seed-data.js with {len(final_data)} entries.")

if __name__ == "__main__":
    run()
